#!/usr/bin/env python3
"""
総務省『全国地方公共団体コード』Excel ファイルを JSON snapshot に変換。

POLICIES §10「cityCode 一次資料準拠ルール」のための一次資料取り込みスクリプト。

Usage:
  python3 scripts/parse-soumu-xlsx.py [input.xls] [output.json]

  デフォルト:
    input  = data/reference/soumu-jichitai-codes-2026-05-11.xls
    output = data/reference/soumu-jichitai-codes.json

総務省Excelの想定フォーマット:
  - 1行目: ヘッダー（団体コード | 都道府県名（漢字）| 市区町村名（漢字）| 都道府県名（カナ）| 市区町村名（カナ））
  - 2行目以降: データ。市区町村名が空のものは都道府県のみ行
  - 団体コードは6桁（5桁地域コード + 1桁チェックデジット）

出力 JSON schema:
  {
    "source": URL,
    "sourceTitle": str,
    "fetchedAt": "YYYY-MM-DD",
    "partialCoverage": false,
    "entries": [
      { "code6": "012025", "code5": "01202", "checkDigit": "5",
        "prefecture": "北海道", "cityName": "函館市",
        "prefectureKana": "ﾎｯｶｲﾄﾞｳ", "cityKana": "ﾊｺﾀﾞﾃｼ",
        "isPrefecture": false },
      ...
    ]
  }
"""

import json, sys, datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_IN = ROOT / "data" / "reference" / "soumu-jichitai-codes-2026-05-11.xls"
DEFAULT_OUT = ROOT / "data" / "reference" / "soumu-jichitai-codes.json"

def normalize_code(raw):
    """6桁にゼロパディング（int で来る場合・str の場合の両対応）"""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # 数字のみ抽出（念のため）
    s = ''.join(c for c in s if c.isdigit())
    if not s:
        return None
    return s.zfill(6)

def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_IN
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not src.exists():
        print(f"❌ 入力ファイル不在: {src}", file=sys.stderr)
        sys.exit(1)

    print(f"読込: {src}", file=sys.stderr)
    wb = load_workbook(src, data_only=True)
    # 想定: 最初のシートにデータがある
    sheet_names = wb.sheetnames
    print(f"  シート: {sheet_names}", file=sys.stderr)
    ws = wb[sheet_names[0]]

    entries = []
    seen_codes = set()
    skipped_no_code = 0
    duplicates = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            # ヘッダー行 — 確認のみ
            print(f"  ヘッダー: {row[:5]}", file=sys.stderr)
            continue
        if not row or all(c is None for c in row):
            continue

        code6 = normalize_code(row[0] if len(row) > 0 else None)
        if not code6:
            skipped_no_code += 1
            continue
        pref = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        city = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else "")
        pref_kana = (str(row[3]).strip() if len(row) > 3 and row[3] is not None else "")
        city_kana = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else "")

        if code6 in seen_codes:
            duplicates += 1
            continue
        seen_codes.add(code6)

        entry = {
            "code6": code6,
            "code5": code6[:5],
            "checkDigit": code6[5],
            "prefecture": pref,
        }
        if city:
            entry["cityName"] = city
            entry["isPrefecture"] = False
        else:
            entry["isPrefecture"] = True
        if pref_kana:
            entry["prefectureKana"] = pref_kana
        if city and city_kana:
            entry["cityKana"] = city_kana

        entries.append(entry)

    municipalities = sum(1 for e in entries if not e.get("isPrefecture"))
    prefectures = sum(1 for e in entries if e.get("isPrefecture"))

    doc = {
        "source": "https://www.soumu.go.jp/main_content/000925835.xls",
        "sourceTitle": "都道府県コード及び市区町村コード（総務省・令和6年1月1日更新）",
        "sourceFile": src.name,
        "fetchedAt": "2026-05-11",
        "parsedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "fetchMethod": "openpyxl による XLSX 解析（OOXML 形式）",
        "partialCoverage": False,
        "coverageNote": "総務省 .xls 全件取り込み完了。本ファイルが cityCode の一次正準ソース。",
        "totalEntries": len(entries),
        "municipalityEntries": municipalities,
        "prefectureEntries": prefectures,
        "entries": entries
    }

    with open(dst, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"✅ 書出: {dst}", file=sys.stderr)
    print(f"   合計: {len(entries)} エントリ", file=sys.stderr)
    print(f"     自治体: {municipalities}", file=sys.stderr)
    print(f"     都道府県のみ: {prefectures}", file=sys.stderr)
    if skipped_no_code:
        print(f"   ⚠️  cityCode 不明でスキップ: {skipped_no_code} 行", file=sys.stderr)
    if duplicates:
        print(f"   ⚠️  重複 cityCode（先勝ち）: {duplicates} 行", file=sys.stderr)

if __name__ == "__main__":
    main()
